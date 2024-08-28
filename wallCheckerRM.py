# This code is part of the Master Thesis of Jean van der Meer presented to the Eindhoven University of Technology
# Room Mode version of the Wall Checker. The Excel workbook generating function also has some changes to work with Room Mode.
# The Room Mode functionality is defined here, to check whether the ifc walls that ought to be checked belong 
# to the scanned area (that can be checked) or not.

import numpy as np
import alphashape
from shapely.geometry import Polygon, Point
import pandas as pd
import os
from scipy.spatial import Delaunay
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d.art3d import Poly3DCollection

# Function to read point cloud from a file and use it to later find the volume that bounds the point cloud
def read_point_cloud2(file_path):
    points = []
    with open(file_path, 'r') as f:
        for line in f:
            parts = line.split()
            if len(parts) >= 3:  # Ensure there are at least 3 components
                # for every split line divide the first 3 values as x, y and z and later assign the tuple of them as a point into a list of points
                x, y, z = map(float, parts[:3])
                points.append((x, y, z))
    return np.array(points)

# Function to downsample the point cloud using voxel grid filtering
# Some point clouds have a very high point density, which makes computation of the bounding volume of the point cloud slow. The same computation can be 
# achieved at a lesser density of points, so that is done here.
def voxel_grid_downsample(points, voxel_size):
    coords_min = np.min(points, axis=0)
    coords_max = np.max(points, axis=0)
    dims = np.ceil((coords_max - coords_min) / voxel_size).astype(int)
    
    # Calculate voxel indices
    indices = np.floor((points - coords_min) / voxel_size).astype(int)
    
    # Use np.unique to find unique voxel indices and corresponding points
    _, unique_indices = np.unique(indices, axis=0, return_index=True)
    
    return points[unique_indices]

# Function to compute the 2D concave hull, buffer it, and extrude it
# Instead of calculating a bounding volume (hull) around the entire point cloud, which would again be very computationally expensive,
# the point cloud is "flattened", that is, only the x and y coordinates are used to calculate a bounding area of the scanned area.
# because people mostly only walk in 2 dimensions in indoor environments, the data for most indoor environment layouts is expressed in x, y, 
# and can be mostly just extruded in the z direction, with ceilings usually having similar heights. Skipping the computation of this third dimension 
# can speed enormously the computation process, from more than 40 minutes (sometimes even crashes) to around a minute or less for the largest
# project used. In some very specific cases where height of ceilings change a lot, as in floors where one area has a double height ceiling and the
# other area does not, and only that floor is scanned and other areas shouldn't be checked or compared to the IFC data of non scanned areas of
# the next floor, a solution with voxelization of the point cloud data to create a voxelized volume could speed up the computation and still
# take variations in the z coordinate into account.
# A buffer is made around the calculated volume, to find wall starting and end points that might have fallen just outside the scanned area due
# to imprecisions of the scanning process of discrepancies from as-designed and as-built measurements. Alpha is a factor that determines how 
# small are the concavities that the algorithm should look for when creating a volume that bounds the points
def compute_2d_concave_hull_and_extrude(points, alpha=1.0, buffer_size=0.4):
    # Project points onto the XY plane (flatten the Z coordinate)
    voxel_size = 0.5
    points = voxel_grid_downsample(points, voxel_size)
    # just the downsampled x and y coordinates of points are used to make the bounding area
    points_2d = points[:, :2]

    # Compute the 2D concave hull
    hull = alphashape.alphashape(points_2d, alpha)
    hull_polygon = Polygon(hull.exterior.coords)
    
    # Apply buffer to the hull 
    expanded_hull_polygon = hull_polygon.buffer(buffer_size)
    expanded_hull_points = np.array(expanded_hull_polygon.exterior.coords)
    
    # Get the Z range for extrusion
    z_min = np.min(points[:, 2]) - 0.3 # make sure that the base points of walls are checked even if the scan is a bit higher than the ifc floor
    z_max = np.max(points[:, 2]) - 0.3 # make sure that the walls of the next floor are not considered as belonging to the scanned floor

    # Plot the 3D extruded concave hull
    fig = plt.figure()
    ax = fig.add_subplot(111, projection='3d')

    # Create the sides of the extruded shape
    for i in range(len(expanded_hull_points) - 1):
        x = [expanded_hull_points[i, 0], expanded_hull_points[i + 1, 0], expanded_hull_points[i + 1, 0], expanded_hull_points[i, 0]]
        y = [expanded_hull_points[i, 1], expanded_hull_points[i + 1, 1], expanded_hull_points[i + 1, 1], expanded_hull_points[i, 1]]
        z = [z_min, z_min, z_max, z_max]
        ax.add_collection3d(Poly3DCollection([list(zip(x, y, z))], color='cyan', alpha=0.5))

    # Create the top and bottom faces
    for z in [z_min, z_max]:
        x = expanded_hull_points[:, 0]
        y = expanded_hull_points[:, 1]
        ax.add_collection3d(Poly3DCollection([list(zip(x, y, [z]*len(x)))], color='cyan', alpha=0.5))

    ax.set_xlabel('X')
    ax.set_ylabel('Y')
    ax.set_zlabel('Z')

    # Set default axis limits for visualization
    ax.set_xlim(-25.0, 25.0)
    ax.set_ylim(-25.0, 25.0)
    ax.set_zlim(-12.0, 12.0)
    # visualize bounding hull
    plt.show()

    return expanded_hull_polygon

# Function to check if a point is within the alpha hull
def is_within_alpha_hull(point, alpha_hull_polygon, buffer_size=0.55):
    point_3d = Point(point[:3])
    buffered_polygon = alpha_hull_polygon.buffer(buffer_size)
    return buffered_polygon.contains(point_3d)

# Function to extract a wall's start and end points
def extrPoints(wall):
    local_placement = wall.ObjectPlacement
    if local_placement:
        location = local_placement.RelativePlacement.Location.Coordinates
        # new walls created in the tool may not have an elevation yet at some checkpoints
        if wall.ContainedInStructure[0].RelatingStructure.Elevation:
            z_coord = wall.ContainedInStructure[0].RelatingStructure.Elevation
        else:
            # newly created walls start with a global z height in their "location" z coordinate and later get assigned 
            # to a floor and don't need a global z height anymore (it is corrected to relative z), as their height 
            # is then referenced by the floor
            z_coord = local_placement.RelativePlacement.Location.Coordinates[2]
        # the length of the wall helps finding the end point of it, that is implicit in IFC. Points[1] is the second index,
        # so the end point, and the length is given in the x axis, the first axis, so Coordinates[0]    
        lengthW = wall.Representation.Representations[0].Items[0].Points[1].Coordinates[0]
        # For some horizontal walls RefDirection can be None, so they will not have an individual DirectionRatio as they
        # already follow the direction of the floor they are located in
        if local_placement.RelativePlacement.RefDirection is not None:
            axis2 = local_placement.RelativePlacement.RefDirection.DirectionRatios
            axis1 = local_placement.RelativePlacement.Axis.DirectionRatios
            # A direction ratio of (-1,0,0) means the wall is horizontal (x axis = 1) but the point stored as 
            # location point of the wall is at the end of the wall, when looking at global coordinates
            if axis1 == (0,0,1) and axis2 == (-1,0,0):
                end_coordinate = ((location[0] - lengthW), location[1], location[2])
            # A direction ratio of (0,1,0) and (0,-1,0) mean the wall is vertical (y axis = 1 or -1) for (0,1,0) the 
            # location point is in a lower global y value compared to the end of the wall, and for (0,-1,0) the location/base
            # point of the wall is located on a higher global y value and the end point at a lower y value than the base point
            elif axis1 == (0,0,1) and axis2 == (0,1,0):
                end_coordinate = (location[0], (location[1] + lengthW), location[2])
            elif axis1 == (0,0,1) and axis2 == (0,-1,0):
                end_coordinate = (location[0], (location[1] - lengthW), location[2])
        else:
            end_coordinate = (location[0] + wall.Representation.Representations[0].Items[0].Points[1].Coordinates[0], location[1], location[2])
        return (location[0], location[1], z_coord), (end_coordinate[0], end_coordinate[1], z_coord)
    return None

# Function to process walls in Room Mode
def process_seg_wallsRM(files2):
    data_dict = {}
    wall_dict = {}

    for file in files2:
        # Read the point clouds and store them in data_dict
        data_dict[file] = pd.read_csv(file, sep=' ', header=None, names=['X', 'Y', 'Z', 'R', 'G', 'B'])

    for file, data in data_dict.items():
        x = data['X'].values
        y = data['Y'].values
        z = data['Z'].values
        x_diff = np.abs(x.max() - x.min())
        y_diff = np.abs(y.max() - y.min())

        wall_name = os.path.splitext(os.path.basename(file))[0] # Extract the base name without the extension, as the walls are handled internally as wall1, wall2 etc and not wall1.txt, wall2.txt etc

        if x_diff <= 0.78 and x_diff < y_diff:
            # The threshold here at room mode is much higher than at the older wallChecker. that is mainly due to walls tested at bigger projects being often wider.
            # This threshold difference at x_diff can be changed, it is mainly placed to estimate the thickness of a wall, and assumes that if the x coordinates 
            # of a wall stay constant and within a range that can be considered the thickness of a wall, and the y coordinates change a lot, the wall is vertical
            # seen from a plan view, that is, longitudinally grows along the y axis
            wall_dict[wall_name] = {
                'type': 'vertical',
                'base point': (float(x.mean()), float(y.min()), float(z.min())),
                'end point': (float(x.mean()), float(y.max()), float(z.min())),
                'height': float(z.max()-z.min()),
                'thickness': float(x.max()-x.min()),
                'length': float((y.max() - y.min()))
            }
        elif y_diff <= 0.78 and y_diff < x_diff:
            wall_dict[wall_name] = {
                'type': 'horizontal',
                'base point': (float(x.min()), float(y.mean()), float(z.min())),
                'end point': (float(x.max()), float(y.mean()), float(z.min())),
                'height': float(z.max() - z.min()),
                'thickness': float(y.max()-y.min()),
                'length': float((x.max() - x.min()))
            }
        else:
            # Here is where walls that follow a non-manhattan world assumption can be handled. Their start and end points can be found by a linear approximation of the
            # x and y coordinates of the segmented wall
            wall_dict[wall_name] = {
                'type': 'diagonal wall',
                'base point': (),
                'end point': (),
                'height': float(z.max()-z.min())
            }

    return wall_dict

# Function to match walls in Room Mode
def wallMatcherRM(model, wall_dict, alpha_hull, buffer_size=0.55):
    point_cloud_walls_matched = []
    ifc_walls_matched = []
    ifc_walls_to_delete = []

    for wall in wall_dict:
        wall_matched = False
        for ifc_wall in model.by_type("IfcWallStandardCase"):
            if ifc_wall.Representation.Representations[1].Items[0].SweptArea.is_a('IfcRectangleProfileDef'):
                start_point, end_point = extrPoints(ifc_wall)
                # Here a dynamic threshold is created and used to match IFC and point cloud walls. Because the distance between the start point of a wall
                # in an IFC file and the start point of the same wall in a point cloud can be quite considerable, taken deviations of measeurement, as-designed
                # vs as-built differences, and differences in how wall connections and starting points are defined (discussed in the report), this distance
                # can be of around a metre or even a bit more for walls that are 70 cm thick, for instance. It would not be realistic to use a thresholf of
                # more than 70 cm for 10cm thick walls however, so a minimum threshold is defined, that is used for most thin walls, and a threshold
                # proportional to wall thickness is used for walls that are very thick. To save space the threshold is named dth (Dynamic ThresHold).
                
                ifc_wall_dim_y = ifc_wall.Representation.Representations[1].Items[0].SweptArea.YDim
                dth = max(0.65, 2.5*ifc_wall_dim_y)
                # The big OR conditional below defines whether walls are matched start to start and end to end OR start to end and end to start, because
                # IFC can structure the global coordinates of their wall starts and ends in a counterintuitive direction, which is converted in IFC
                # by a Reference Direction. Because point clouds always work in the same coordinate system we need to check for both possibilities though
                # to make sure a match is fully checked
                if is_within_alpha_hull(start_point, alpha_hull, buffer_size) and is_within_alpha_hull(end_point, alpha_hull, buffer_size):
                    if (
                        (
                            (wall_dict[wall]['base point'][0] < (start_point[0] + dth) and wall_dict[wall]['base point'][0] > (start_point[0] - dth)) and 
                            (wall_dict[wall]['base point'][1] < (start_point[1] + dth) and wall_dict[wall]['base point'][1] > (start_point[1] - dth)) and 
                            (wall_dict[wall]['end point'][0] < (end_point[0] + dth) and wall_dict[wall]['end point'][0] > (end_point[0] - dth)) and 
                            (wall_dict[wall]['end point'][1] < (end_point[1] + dth) and wall_dict[wall]['end point'][1] > (end_point[1] - dth))
                        ) or 
                        (
                            (wall_dict[wall]['base point'][0] < (end_point[0] + dth) and wall_dict[wall]['base point'][0] > (end_point[0] - dth)) and 
                            (wall_dict[wall]['base point'][1] < (end_point[1] + dth) and wall_dict[wall]['base point'][1] > (end_point[1] - dth)) and 
                            (wall_dict[wall]['end point'][0] < (start_point[0] + dth) and wall_dict[wall]['end point'][0] > (start_point[0] - dth)) and 
                            (wall_dict[wall]['end point'][1] < (start_point[1] + dth) and wall_dict[wall]['end point'][1] > (start_point[1] - dth))
                        )
                    ):
                        print(f'Wall {wall} at the point cloud has matched wall {ifc_wall.GlobalId} at the IFC file')
                        wall_matched = True
                        point_cloud_walls_matched.append(wall)
                        ifc_walls_matched.append(ifc_wall.GlobalId)
        if not wall_matched:
            print(f'Wall {wall} at the point cloud did not find a match in the IFC file. It needs to be modeled in the IFC file.')
                    

    for ifc_wall in model.by_type("IfcWallStandardCase"):
        start_point, end_point = extrPoints(ifc_wall)
        if is_within_alpha_hull(start_point, alpha_hull, buffer_size) and is_within_alpha_hull(end_point, alpha_hull, buffer_size):
            if ifc_wall.GlobalId not in ifc_walls_matched:
                print(f'Wall {ifc_wall.GlobalId} in the IFC file did not find a match in the point cloud. It needs to be deleted from the IFC file.')
                ifc_walls_to_delete.append(ifc_wall.GlobalId)
    # here a direct list of walls to be deleted is created, as only walls that are not checked withing the scanned area should be removed,
    # instead of walls of the entire model that don't find a match with point cloud data
    return ifc_walls_matched, point_cloud_walls_matched, ifc_walls_to_delete



###########################
# Export results to Excel #
# Here, an excel file is generated that has as-designed IFC walls, their global ids and names to say which ones 
# are present in the real building, conforming to the design concept of the building, and which ones are not.
# furthermore, walls in the as-built/as-is condition of the building that are present in the point cloud but
# were not present in the IFC file have their global location outputted so a new wall can be drawn automatically
# or by a modeller, if necessary, and checked.

def resultsExcel(model, wall_dict, ifc_walls_matched, point_cloud_walls_matched, alpha_hull, buffer_size=0.55):
    # from inter5 import model
    import openpyxl
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter

    # Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active

    # Set column headers
    ws['A1'] = 'IFC Wall Name'
    ws['B1'] = 'IFC Wall GUID'
    ws['C1'] = 'Status'

    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 40

    # Set initial row index
    row_index = 2

    # Iterate over IFC walls
    for ifc_wall in model.by_type("IfcWallStandardCase"):
        if ifc_wall.Representation.Representations[1].Items[0].SweptArea.is_a('IfcRectangleProfileDef'):
                start_point, end_point = extrPoints(ifc_wall)
                # only walls within the alpha hull should be marked as not checked, so the test is done again
                if is_within_alpha_hull(start_point, alpha_hull, buffer_size) and is_within_alpha_hull(end_point, alpha_hull, buffer_size):

                    if ifc_wall.GlobalId in ifc_walls_matched:
                        # Match found, set status and fill cell with green color for IFC walls that found a match
                        ws.cell(row=row_index, column=1, value=ifc_wall.Name)
                        ws.cell(row=row_index, column=2, value=ifc_wall.GlobalId)
                        ws.cell(row=row_index, column=3, value=f'Matched with {point_cloud_walls_matched[ifc_walls_matched.index(ifc_wall.GlobalId)]}')
                        ws.cell(row=row_index, column=3).fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                    else:
                        # No match found, set status and fill cell with red color
                        ws.cell(row=row_index, column=1, value=ifc_wall.Name)
                        ws.cell(row=row_index, column=2, value=ifc_wall.GlobalId)
                        ws.cell(row=row_index, column=3, value='No match found, delete wall')
                        ws.cell(row=row_index, column=3).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    
                    row_index += 1

    # Add table of point cloud walls that are matched or not
    row_index += 2
    ws.cell(row=row_index, column=1, value='Point Cloud Wall Name')
    ws.cell(row=row_index, column=2, value='Matched IFC Wall')
    ws.cell(row=row_index, column=3, value='Coordinates to build new wall, if needed')

    row_index += 1

    for wall in wall_dict:
        ws.cell(row=row_index, column=1, value=wall)
        if wall in point_cloud_walls_matched:
            ws.cell(row=row_index, column=2, value=ifc_walls_matched[point_cloud_walls_matched.index(wall)])
        else:
            ws.cell(row=row_index, column=2, value='No match found')
            ws.cell(row=row_index, column=2).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        start_point = [round(coord, 6) for coord in wall_dict[wall]["base point"]]
        end_point = [round(coord, 6) for coord in wall_dict[wall]["end point"]]
        ws.cell(row=row_index, column=3, value=f'Start: {start_point}, End: {end_point}')
        
        row_index += 1

    # Set number format for coordinate columns
    for col in ['C']:
        for row in range(2, row_index):
            cell = ws[f'{col}{row}']
            cell.number_format = '0.000000'

    # Save the workbook
    wb.save('wall_matching_results.xlsx')

