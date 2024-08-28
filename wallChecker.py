# This code is part of the Master Thesis of Jean van der Meer presented to the Eindhoven University of Technology
import ifcopenshell
import ifcopenshell.geom
import numpy as np
import math
import os
import pandas as pd
import glob


# here the IFC file of the as-designed project is opened, to be compared with the point clouds,
# have its outdated elements pointed out in a report, and be (later, in a different import) updated into a new IFC file

########################################################################################
# First step: find beginning and end coordinates of the walls in the Point cloud file  #
# It is necessary to classify them in horizontal or vertical in order to be able to    #
# know which maximum and minimum values to seek and how they will translate into start #
# and end coordinates of the wall                                                      #
########################################################################################

# Here all walls that were segmented are read into the script and saved in a dictionary
# the segmentation method assumes that each wall is in a separate text file (both sides of the wall).
# some other methodologies have a point cloud with one of the values per line pointing out that 
# that specific point is a point that belongs to a given wall, or that it belongs to some other type
# of building element. It is possible to adapt the code into such a methodology, but that is not the 
#  methodology currently adopted. Some reasons for that are discussed in the thesis report

# point cloud walls are assumed to be named wall1, wall2, wall3, etc for each wall and each text file
# here, point clouds are used in txt format following the xyz standard, with no header in the file
# a point cloud in xyz can be used by changing the (prefix + '*.txt') below to (prefix + '*.xyz')

# Because the only information relevant for the heuristics used here is the x,y,z geometric information, 
# an accompanying file named pcdSimplifier2.py is added to this repository that can clean point cloud files
# from other data and make the files lighter

# In the interface edition of the code, walls are automatically named in the right standard when loaded from
# a folder 

import glob
import pandas as pd
import numpy as np
import os

def process_seg_walls(files2):
    data_dict = {}
    wall_dict = {}

    for file in files2:
        # Read the point clouds and store them in data_dict
        data_dict[file] = pd.read_csv(file, sep=' ', header=None, names=['X', 'Y', 'Z', 'R', 'G', 'B'])

    for file, data in data_dict.items():
        x = data['X'].values
        y = data['Y'].values
        z = data['Z'].values
        x_diff = np.abs(x.max() - x.min())
        y_diff = np.abs(y.max() - y.min())

        wall_name = os.path.splitext(os.path.basename(file))[0] # Extract the base name without the extension, as the walls are handled internally as wall1, wall2 etc and not wall1.txt, wall2.txt etc

        if x_diff <= 0.22:
            # this threshold difference at x_diff can be changed, it is mainly placed to estimate the thickness of a wall, and assumes that if the x coordinates 
            # of a wall stay constant and within a range that can be considered the thickness of a wall, and the y coordinates change a lot, the wall is vertical
            # seen from a plan view, that is, longitudinally grows along the y axis
            wall_dict[wall_name] = {
                'type': 'vertical',
                'base point': (float(x.mean()), float(y.min()), float(z.min())),
                'end point': (float(x.mean()), float(y.max()), float(z.min())),
                'height': float(z.max()-z.min()),
                'thickness': float(x.max()-x.min()),
                'length': float((y.max() - y.min()))
            }
        elif y_diff <= 0.22:
            wall_dict[wall_name] = {
                'type': 'horizontal',
                'base point': (float(x.min()), float(y.mean()), float(z.min())),
                'end point': (float(x.max()), float(y.mean()), float(z.min())),
                'height': float(z.max() - z.min()),
                'thickness': float(y.max()-y.min()),
                'length': float((x.max() - x.min()))
            }
        else:
            # Here is where walls that follow a non-manhattan world assumption can be handled. Their start and end points can be found by a linear approximation of the
            # x and y coordinates of the segmented wall
            wall_dict[wall_name] = {
                'type': 'diagonal wall',
                'base point': (),
                'end point': (),
                'height': float(z.max()-z.min())
            }

    return wall_dict


# print(wall_dict['wall1']['type'])

#############################################################################################
# Second step: find beginning and end coordinates of the walls in the as-designed IFC file  #
# Depending on the RefDirection that orients the starting point of a wall, which can be     #
# random (starting at beginning or end), the end coordinate will be found, based on the     #
# shape representation polyline that gives the length of the wall; the local placement, that#
# gives the start coordinate of the wall; and the axis and Reference Direction of           #            
# the wall that transform the local axis into the global axis                               #
#############################################################################################

import ifcopenshell
import ifcopenshell.util.placement

# Function to extract a wall's start and end points
def extrPoints(wall):
    local_placement = wall.ObjectPlacement
    if local_placement:
        location = local_placement.RelativePlacement.Location.Coordinates
        # new walls created in the tool may not have an elevation yet at some checkpoints
        if wall.ContainedInStructure[0].RelatingStructure.Elevation:
            z_coord = wall.ContainedInStructure[0].RelatingStructure.Elevation
        else:
            # newly created walls start with a global z height in their "location" z coordinate and later get assigned 
            # to a floor and don't need a global z height anymore (it is corrected to relative z), as their height 
            # is then referenced by the floor
            z_coord = local_placement.RelativePlacement.Location.Coordinates[2]
        # the length of the wall helps finding the end point of it, that is implicit in IFC. Points[1] is the second index,
        # so the end point, and the length is given in the x axis, the first axis, so Coordinates[0]    
        lengthW = wall.Representation.Representations[0].Items[0].Points[1].Coordinates[0]
        # For some horizontal walls RefDirection can be None, so they will not have an individual DirectionRatio as they
        # already follow the direction of the floor they are located in
        if local_placement.RelativePlacement.RefDirection is not None:
            axis2 = local_placement.RelativePlacement.RefDirection.DirectionRatios
            axis1 = local_placement.RelativePlacement.Axis.DirectionRatios
            # A direction ratio of (-1,0,0) means the wall is horizontal (x axis = 1) but the point stored as 
            # location point of the wall is at the end of the wall, when looking at global coordinates
            if axis1 == (0,0,1) and axis2 == (-1,0,0):
                end_coordinate = ((location[0] - lengthW), location[1], location[2])
            # A direction ratio of (0,1,0) and (0,-1,0) mean the wall is vertical (y axis = 1 or -1) for (0,1,0) the 
            # location point is in a lower global y value compared to the end of the wall, and for (0,-1,0) the location/base
            # point of the wall is located on a higher global y value and the end point at a lower y value than the base point
            elif axis1 == (0,0,1) and axis2 == (0,1,0):
                end_coordinate = (location[0], (location[1] + lengthW), location[2])
            elif axis1 == (0,0,1) and axis2 == (0,-1,0):
                end_coordinate = (location[0], (location[1] - lengthW), location[2])
        else:
            end_coordinate = (location[0] + wall.Representation.Representations[0].Items[0].Points[1].Coordinates[0], location[1], location[2])
        return (location[0], location[1], z_coord), (end_coordinate[0], end_coordinate[1], z_coord)
    return None


##############################################################################################
# Third step: Iterate over walls in the point cloud and check whether any of the walls       #
# in the IFC file matches this point cloud wall. This is done by checking whether there is   #
# an absolute difference smaller than e.g. ~0.22m for the Y and X coordinates, of either     #
# Base Point of Point cloud and base point of ifc wall, AND end point of point cloud and end #
# point of ifc wall, OR a match within 0.22m of the base point of point cloud and end point  #
# of ifc wall, AND end point of point cloud and base point of ifc wall, because the ifc      #
# file does not necessarily consider the wall as always starting from left to right and      #
# bottom to top as the point cloud files do.                                                 #            
##############################################################################################
# in Room Mode this 0.22 threshold is higher and is also a dynamic threshold, depending on the thickness of the walls being compared and a minimum value
def wallMatcher(model, wall_dict):
    # from inter5 import model
    point_cloud_walls_matched = []
    ifc_walls_matched = []

    for wall in wall_dict:
        wall_matched = False
        for ifc_wall in model.by_type("IfcWallStandardCase"):
            if (
                (
                    (wall_dict[wall]['base point'][0] < (extrPoints(ifc_wall)[0][0] + 0.22) and wall_dict[wall]['base point'][0] > (extrPoints(ifc_wall)[0][0] - 0.22)) and 
                    (wall_dict[wall]['base point'][1] < (extrPoints(ifc_wall)[0][1] + 0.22) and wall_dict[wall]['base point'][1] > (extrPoints(ifc_wall)[0][1] - 0.22)) and 
                    (
                        (wall_dict[wall]['end point'][0] < (extrPoints(ifc_wall)[1][0] + 0.22) and wall_dict[wall]['end point'][0] > (extrPoints(ifc_wall)[1][0] - 0.22)) and 
                        (wall_dict[wall]['end point'][1] < (extrPoints(ifc_wall)[1][1] + 0.22) and wall_dict[wall]['end point'][1] > (extrPoints(ifc_wall)[1][1] - 0.22))
                    )
                ) or 
                (
                    (
                        (wall_dict[wall]['base point'][0] < (extrPoints(ifc_wall)[1][0] + 0.22) and wall_dict[wall]['base point'][0] > (extrPoints(ifc_wall)[1][0] - 0.22)) and 
                        (wall_dict[wall]['base point'][1] < (extrPoints(ifc_wall)[1][1] + 0.22) and wall_dict[wall]['base point'][1] > (extrPoints(ifc_wall)[1][1] - 0.22)) and 
                        (
                            (wall_dict[wall]['end point'][0] < (extrPoints(ifc_wall)[0][0] + 0.22) and wall_dict[wall]['end point'][0] > (extrPoints(ifc_wall)[0][0] - 0.22)) and 
                            (wall_dict[wall]['end point'][1] < (extrPoints(ifc_wall)[0][1] + 0.22) and wall_dict[wall]['end point'][1] > (extrPoints(ifc_wall)[0][1] - 0.22))
                        )
                    )
                )
            ):
                print(f'Wall {wall} at the point cloud has matched wall {ifc_wall.GlobalId} at the IFC file')
                wall_matched = True
                point_cloud_walls_matched.append(wall)
                ifc_walls_matched.append(ifc_wall.GlobalId)
        
        if not wall_matched:
            # The walls present in the point cloud (as-is / as-built) that were not matched with the IFC model are
            # new walls or walls with a new configuration, that needs to be modelled. A report is made, and later in 
            # the code, they are updated into the IFC file for some of the use cases
            print(f'Wall {wall} at the point cloud did not find a match in the IFC file. It needs to be modeled in the IFC file.')

    for ifc_wall in model.by_type("IfcWallStandardCase"):
        if ifc_wall.GlobalId not in ifc_walls_matched:
            # Walls present in the as-designed model, but that are not found in the current building, should be deleted from the IFC project
            print(f'Wall {ifc_wall.GlobalId} in the IFC file did not find a match in the point cloud. It needs to be deleted from the IFC file.')
    return ifc_walls_matched, point_cloud_walls_matched


###########################
# Export results to Excel #
# Here, an excel file is generated that has as-designed IFC walls, their global ids and names to say which ones 
# are present in the real building, conforming to the design concept of the building, and which ones are not.
# furthermore, walls in the as-built/as-is condition of the building that are present in the point cloud but
# were not present in the IFC file have their global location outputted so a new wall can be drawn automatically,
# or by a modeller, if necessary, and checked.

def resultsExcel(model, wall_dict, ifc_walls_matched, point_cloud_walls_matched):
    
    import openpyxl
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter

    # Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active

    # Set column headers
    ws['A1'] = 'IFC Wall Name'
    ws['B1'] = 'IFC Wall GUID'
    ws['C1'] = 'Status'

    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 40

    # Set initial row index
    row_index = 2

    # Iterate over IFC walls
    for ifc_wall in model.by_type("IfcWallStandardCase"):
        if ifc_wall.GlobalId in ifc_walls_matched:
            # Match found, set status and fill cell with green color for IFC walls that found a match
            ws.cell(row=row_index, column=1, value=ifc_wall.Name)
            ws.cell(row=row_index, column=2, value=ifc_wall.GlobalId)
            ws.cell(row=row_index, column=3, value=f'Matched with {point_cloud_walls_matched[ifc_walls_matched.index(ifc_wall.GlobalId)]}')
            ws.cell(row=row_index, column=3).fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        else:
            # No match found, set status and fill cell with red color
            ws.cell(row=row_index, column=1, value=ifc_wall.Name)
            ws.cell(row=row_index, column=2, value=ifc_wall.GlobalId)
            ws.cell(row=row_index, column=3, value='No match found, delete wall')
            ws.cell(row=row_index, column=3).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        
        row_index += 1

    # Add table of point cloud walls that are matched or not
    row_index += 2
    ws.cell(row=row_index, column=1, value='Point Cloud Wall Name')
    ws.cell(row=row_index, column=2, value='Matched IFC Wall')
    ws.cell(row=row_index, column=3, value='Coordinates to build new wall, if needed')

    row_index += 1

    for wall in wall_dict:
        ws.cell(row=row_index, column=1, value=wall)
        if wall in point_cloud_walls_matched:
            ws.cell(row=row_index, column=2, value=ifc_walls_matched[point_cloud_walls_matched.index(wall)])
        else:
            ws.cell(row=row_index, column=2, value='No match found')
            ws.cell(row=row_index, column=2).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        # round start and end point coordinates to limit cell size
        start_point = [round(coord, 6) for coord in wall_dict[wall]["base point"]]
        end_point = [round(coord, 6) for coord in wall_dict[wall]["end point"]]
        ws.cell(row=row_index, column=3, value=f'Start: {start_point}, End: {end_point}')
        
        row_index += 1

    # Set number format for coordinate columns
    for col in ['C']:
        for row in range(2, row_index):
            cell = ws[f'{col}{row}']
            cell.number_format = '0.000000'

    # Save the workbook
    wb.save('wall_matching_results.xlsx')

